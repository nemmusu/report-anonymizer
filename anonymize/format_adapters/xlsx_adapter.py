"""XLSX adapter built on openpyxl.

One segment per cell that holds a string (or a string-like value coerced via
``str()``). Formulas are NOT touched by default (substitutions in formula
strings are opt-in via ``include_formulas=True`` at adapter construction time).
"""
from __future__ import annotations

from pathlib import Path

from .base import (
    FormatAdapter,
    Segment,
    SubstitutionRule,
    WriteEvent,
    WriteReport,
    apply_to_text,
)


class XlsxAdapter(FormatAdapter):
    name = "xlsx"
    extensions = {".xlsx", ".xlsm"}
    mimes = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroenabled.12",
    }

    def __init__(self, include_formulas: bool = False) -> None:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as e:  # pragma: no cover
            raise RuntimeError(
                f"openpyxl is required for the xlsx adapter (pip install openpyxl): {e}"
            )
        self._load = load_workbook
        self.include_formulas = include_formulas

    @staticmethod
    def _is_formula(cell) -> bool:
        v = cell.value
        return isinstance(v, str) and v.startswith("=")

    def extract(self, path: Path) -> list[Segment]:
        wb = self._load(str(path), data_only=False, read_only=False)
        out: list[Segment] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if v is None:
                        continue
                    if not isinstance(v, str):
                        continue
                    if not self.include_formulas and self._is_formula(cell):
                        continue
                    out.append(
                        Segment(
                            seg_id=f"{ws.title}!{cell.coordinate}",
                            text=v,
                            meta={"sheet": ws.title, "coord": cell.coordinate},
                        )
                    )
        wb.close()
        return out

    def write(
        self,
        src_path: Path,
        dst_path: Path,
        substitutions: list[SubstitutionRule],
    ) -> WriteReport:
        wb = self._load(str(src_path), data_only=False, read_only=False)
        events: list[WriteEvent] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if not isinstance(v, str):
                        continue
                    if not self.include_formulas and self._is_formula(cell):
                        continue
                    seg_id = f"{ws.title}!{cell.coordinate}"
                    new_v, ev = apply_to_text(v, substitutions, seg_id=seg_id)
                    if new_v != v:
                        cell.value = new_v
                        events.extend(ev)
        dst_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(dst_path))
        wb.close()
        return WriteReport(file_rel=str(dst_path), events=events)


__all__ = ["XlsxAdapter"]
